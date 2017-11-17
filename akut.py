#!/usr/bin/env python
# Test implementation of a working xlsx2json for the akut data sheet

import json
import os

from openpyxl import load_workbook 

in_file = "akut.xlsx"       # name of the input file

wb_in = load_workbook(filename = in_file)

# grab the active worksheet
ws_in = wb_in.active

measurements = []

# the actual measurements start only here
# as the information is hardcoded anyway, no need to
# make anything configurable..
for i in range(5, ws_in.max_row): 

    measurement = {}

    measurement["analyse"] = ws_in.cell(row=i, column=1).value
    measurement["location"] = ws_in.cell(row=i, column=2).value
    measurement["lat"] = ws_in.cell(row=i, column=3).value
    measurement["long"] = ws_in.cell(row=i, column=4).value
    measurement["author"] = ws_in.cell(row=i, column=5).value
    measurement["timestamp"] = ws_in.cell(row=i, column=6).value

    for k in range(8, ws_in.max_column): 

        vid = ws_in.cell(row=2, column=k).value
        if vid is not None:
            measurement["id"] = ws_in.cell(row=2, column=k).value
            measurement["method"] = ws_in.cell(row=3, column=k).value
            measurement["name"] = ws_in.cell(row=4, column=k).value
            measurement["unit"] = ws_in.cell(row=5, column=k).value
            measurement["value"] = ws_in.cell(row=i, column=k).value

            measurements.append(measurement)

# convert to json and just print
print(json.dumps(measurements))

