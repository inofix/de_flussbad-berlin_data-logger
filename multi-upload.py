#!/usr/bin/python
########################################################################
#
# upload.py:
# A simple script for the data-logger..
# TODO: This is a test implementation, for a stable version we
# need some solution for the pip installed lib, data input verification,
# header checks, and it would be nice to also support direct (form-less)
# uploads for our other sources.
#
########################################################################
#
#  This is Free Software; feel free to redistribute and/or modify it
#  under the terms of the GNU General Public License as published by
#  the Free Software Foundation; version 3 of the License.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  Copyright 2017, Michael Lustenberger <mic@inofix.ch>
#
########################################################################

import os
import cgi
import datetime
import json
import re

# debug {{ #
#import cgitb
#cgitb.enable(display=0, logdir="/tmp/")
# debug }} #

def recalculate_akut(in_file_name):
    from openpyxl import load_workbook
    print('      <p>Starting import:</p>')
    try:
        # get the data from file (no direct method found..)
        wb_in = load_workbook(in_file_name)

        # grab the active worksheet
        ws_in = wb_in.active

        measurements = []

        # the actual measurements start only here
        # as the information is hardcoded anyway, no need to
        # make anything configurable..
        for i in range(6, ws_in.max_row):

            analyse = str(ws_in.cell(row=i, column=1).value)
            location = str(ws_in.cell(row=i, column=2).value)
            # geolocation lat/long are floats
            geo_lat = float(ws_in.cell(row=i, column=3).value)
            geo_long = float(ws_in.cell(row=i, column=4).value)
            author = str(ws_in.cell(row=i, column=5).value)
            # get a timestamp as a string
            t = ws_in.cell(row=i, column=6).value
            if t:
                t = str(t)
                try:
                    # only consider seconds as smallest unit
                    ti = t.index(".")
                    t = t[0:ti]
                except ValueError:
                    pass
            if t:
                try:
                    # is it a timestamp?
                    t_ = datetime.datetime.strptime(t, "%Y-%m-%dT%H:%M:%S")
                    timestamp = t_.isoformat()
                except ValueError:
                    # give it another try..
                    try:
                        t_ = datetime.datetime.strptime(t, "%Y-%m-%d %H:%M:%S")
                        timestamp = t_.isoformat()
                    except ValueError:
                        # last chance, else really fail..
                        t_ = datetime.datetime.strptime(t, "%m/%d/%Y %H:%M:%S")
                        timestamp = t_.isoformat()

            for j in range(8, ws_in.max_column):

                vid = ws_in.cell(row=2, column=j).value
                val = ws_in.cell(row=i, column=j).value
                if vid:
                    name = ws_in.cell(row=4, column=j).value.encode('utf-8')
                    measurement = {
                        "id": str(vid),
                        "method": str(ws_in.cell(row=3, column=j).value),
                        "name": name,
                        "location": location,
                        "name_location": name + " " + location,
                        "unit": ws_in.cell(row=5, column=j).value.encode('utf-8'),
                        "value": val,
                        "timestamp": timestamp,
                        "author": author,
                        "analyse": analyse,
                        "lat": geo_lat,
                        "long": geo_long,
                    }
                    measurements.append(measurement)
                    print("      <p class='detail'>Import OK: " + str(vid) +\
                            " " + measurement["name"] + " = " +\
                            str(val) + ' ' + measurement["unit"] + '</p>\n')

        print('      <p>Import done.</p>')
        return measurements

    except Exception:
        # Quick but Dirty: for now just pretend it to be an input problem...
        raise ValueError('Some value could not be imported, please control your input file...')
### debug ###
#    except Exception:
#        import traceback
#        print('      <p>')
#        print(traceback.format_exc())
#        print('</p>\n')

def store(storage_path, file_prefix='data', do_archive=False):
    timestamp = datetime.datetime.utcnow().replace(microsecond=0).isoformat()
    file_prefix += "-"
    print_h('Current Upload ' + timestamp)

    try:
        data = cgi.FieldStorage()
        if data.has_key('filetype') and data.has_key('file'):
            file_prefix += data['filetype'].value
            if data['filetype'].value == 'inofix':
                j = json.loads(data['file'].value)
                data_file_name = storage_path + '/' + file_prefix +\
                        '-latest.json'
                with open(data_file_name, 'w') as of:
                    of.write(json.dumps(j))
                print_p('File saved.')
                if do_archive:
                    data_file_name = storage_path + '/' + file_prefix +\
                        '-' + timestamp + '.json'
                    with open(data_file_name, 'w') as of:
                        of.write(json.dumps(j))
                    print_p('Archive copy saved.')

            elif data['filetype'].value == 'akut':
                print_p('File received.')
                tmp_file_name = storage_path + '/tmp.xlsx'
                with open(tmp_file_name, 'w') as of:
                    of.write(data['file'].value)
                print_p('File temporarily stored.')
                m = recalculate_akut(tmp_file_name)
                print_p('File converted.')
                data_file_name = storage_path + '/' + file_prefix +\
                        '-latest.json'
                with open(data_file_name, 'w') as of:
                    of.write(json.dumps(m))
                print_p('File saved.')
                if do_archive:
                    data_file_name = storage_path + '/' + file_prefix +\
                        '-' + timestamp + '.json'
                    with open(data_file_name, 'w') as of:
                        of.write(json.dumps(m))
                    print_p('Archive copy saved.')

            else:
                print_p('Unsupported upload format.')
                return
            print_p('Looks all good. Thank you!')
            print_p('This is the link to the <a href="/data-store/' +\
                    file_prefix + '-latest.json">result</a>')
        else:
            print_p('No upload file provided..')
    except Exception as e:
        print_p(e.message)
        print_p('Upload failed. Please try again!')

def print_h(text):
    print('    <h3>' + text + '</h3>\n')

def print_p(text):
    print('    <p>' + text + '</p>\n')

def open_html():
    print('Content-Type: text/html\n\n')
    print('<html>\n')
    print('  <head>\n')
    print('    <meta charset="utf-8" />\n')
    print('  </head>\n')
    print('  <body>\n')

def close_html():
    print('      <p>This is the link (back) to the ')
    print('<a href="/data-store/upload.html">upload form</a></p>\n')
    print('    </body>\n')
    print('</html>\n')

if __name__ == "__main__":

    open_html()

    store("/tmp/", "test", True)

    close_html()
