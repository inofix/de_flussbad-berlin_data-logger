# 
# akut_to_data_manager.py: Read a xlsx-sheet with laboratory values 
# and output a xlsx file readable by INOFIX' data-manager portlet for 
# the Liferay platform.
#
# Created:  2017-11-03 22:42 by Christian Berndt
# Modified: 2017-11-03 22:42 by Christian Berndt
# Version:  1.0.0
# 

import json
import os

from openpyxl import Workbook
from openpyxl import load_workbook 

in_file = "akut.xlsx"       # name of the input file
out_file = "out.xlsx"       # name of the output file

wb_in = load_workbook(filename = in_file)
wb_out = Workbook()

# grab the active worksheet
ws_in = wb_in.active
ws_out = wb_out.active

column_count = ws_in.max_column
column_offset = 6               # first "channel" or "parameter" column

row_count = ws_in.max_row
row_offset = 5                  # first "test" row

output_row = 1                  # idx of the output row



for i in range (row_offset, row_count): 
    
    test_date = ws_in.cell(row=i, column=4).value
    test_location = ws_in.cell(row=i, column=2).value

    for k in range (column_offset, column_count): 

        parameter_method = ws_in.cell(row=1, column=k).value
        parameter_name = ws_in.cell(row=3, column=k).value
        parameter_name = parameter_name + " (" + test_location + ")"

        parameter_unit = ws_in.cell(row=4, column=k).value
        parameter_value = ws_in.cell(row=i, column=k).value

        ws_out.cell(row=output_row, column=1).value=test_date
        ws_out.cell(row=output_row, column=2).value=parameter_name  
        ws_out.cell(row=output_row, column=3).value=parameter_value 
        ws_out.cell(row=output_row, column=4).value=parameter_unit 
        ws_out.cell(row=output_row, column=5).value=parameter_method
        
        output_row = output_row + 1


# Save the file
try: 
    os.remove(out_file)
except OSError:
    pass

wb_out.save(out_file)

print("Wrote data to " + out_file)
